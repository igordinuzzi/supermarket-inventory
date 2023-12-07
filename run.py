from colorama import init, Fore
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Initialize colorama for colored terminal output
init(autoreset=True)

# Constants for the Excel file and sheet names
file_name = 'inventory.xlsx'
sheet_name = 'Sheet 1'

# Function to load the Excel worksheet
def load_worksheet():
    workbook = load_workbook(filename=file_name)
    return workbook[sheet_name]

# Function to save the workbook
def save_workbook(workbook):
    workbook.save(filename=file_name)

# Function to display the current inventory
def display_inventory():
    worksheet = load_worksheet()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        print(Fore.GREEN + f"Item: {row[0]}, Category: {row[1]}, Price: {row[2]}, Expiration Date: {row[3]}, Quantity: {row[4]}")

# Function to add a new item to the inventory
def add_item():
    workbook = load_workbook(filename=file_name)
    worksheet = workbook[sheet_name]

    # Taking inputs from the user, with the option to cancel the operation
    name = input(Fore.WHITE + "Enter item name (or type 'cancel' to exit): ")
    if name.lower() == 'cancel':
        return

    category = input(Fore.WHITE + "Enter item category (or type 'cancel' to exit): ")
    if category.lower() == 'cancel':
        return

    price = input(Fore.WHITE + "Enter item price (or type 'cancel' to exit): ")
    if price.lower() == 'cancel':
        return
    price = float(price)

    expiration_date = input(Fore.WHITE + "Enter expiration date (YYYY-MM-DD) (or type 'cancel' to exit): ")
    if expiration_date.lower() == 'cancel':
        return
    expiration_date = datetime.strptime(expiration_date, '%Y-%m-%d')

    quantity = input(Fore.WHITE + "Enter item quantity (or type 'cancel' to exit): ")
    if quantity.lower() == 'cancel':
        return
    quantity = int(quantity)
    # Append new item to the worksheet and save
    worksheet.append([name, category, price, expiration_date, quantity])
    save_workbook(workbook)
    print(Fore.YELLOW + f"Item '{name}' added with details.")

# Function to update details of an existing item
def update_item():
    workbook = load_workbook(filename=file_name)
    worksheet = workbook[sheet_name]
    # Taking inputs from the user, with the option to cancel the operation
    name = input(Fore.WHITE + "Enter the name of the item to update (or type 'cancel' to exit): ")
    if name.lower() == 'cancel':
        return

    new_name = input(Fore.WHITE + "Enter new name (or press Enter to keep it the same, type 'cancel' to exit): ")
    if new_name.lower() == 'cancel':
        return

    new_category = input(Fore.WHITE + "Enter new category (or press Enter to keep it the same, type 'cancel' to exit): ")
    if new_category.lower() == 'cancel':
        return

    new_price = input(Fore.WHITE + "Enter new price (or press Enter to keep it the same, type 'cancel' to exit): ")
    if new_price.lower() == 'cancel':
        return
    new_price = float(new_price) if new_price else None

    new_expiration_date = input(Fore.WHITE + "Enter new expiration date (YYYY-MM-DD) (or press Enter to keep it the same, type 'cancel' to exit): ")
    if new_expiration_date.lower() == 'cancel':
        return
    new_expiration_date = datetime.strptime(new_expiration_date, '%Y-%m-%d') if new_expiration_date else None

    new_quantity = input(Fore.WHITE + "Enter new quantity (or press Enter to keep it the same, type 'cancel' to exit): ")
    if new_quantity.lower() == 'cancel':
        return
    new_quantity = int(new_quantity) if new_quantity else None
    # Searching and updating the item in the worksheet
    found = False
    for row in worksheet.iter_rows(min_row=2):
        if row[0].value == name:
            row[0].value = new_name or row[0].value
            row[1].value = new_category or row[1].value
            row[2].value = new_price if new_price is not None else row[2].value
            row[3].value = new_expiration_date if new_expiration_date else row[3].value
            row[4].value = new_quantity if new_quantity is not None else row[4].value
            found = True
            break
    # Append new item to the worksheet and save
    save_workbook(workbook)
    # Notify user of update status 
    if found:
        print(Fore.YELLOW + f"Details for '{name}' updated.")
    else:
        print(Fore.RED + f"Item '{name}' not found.")

# Function to delete an item from the inventory
def delete_item():
    workbook = load_workbook(filename=file_name)
    worksheet = workbook[sheet_name]

    # Taking inputs from the user, with the option to cancel the operation
    name = input(Fore.WHITE + "Enter the name of the item to delete (or type 'cancel' to exit): ")
    if name.lower() == 'cancel':
        return

    # Confirmation prompt
    confirm = input(Fore.YELLOW + f"Are you sure you want to delete '{name}'? Type YES to confirm or NO to cancel: ")
    if confirm.lower() != 'yes':
        print(Fore.GREEN + "Deletion cancelled.")
        return

    # Searching and deleting the item in the worksheet
    found = False
    for row in worksheet.iter_rows(min_row=2):
        if row[0].value == name:
            worksheet.delete_rows(row[0].row)
            found = True
            break

    save_workbook(workbook)
    # Notify user of update status 
    if found:
        print(Fore.YELLOW + f"Item '{name}' deleted.")
    else:
        print(Fore.RED + f"Item '{name}' not found.")


# Function to summarize the inventory
def inventory_summary():
    worksheet = load_worksheet()
    # (Code for calculating and displaying the inventory summary)
    total_items = 0
    category_count = {}
    total_quantity = 0
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        total_items += 1
        total_quantity += row[4]
        category_count[row[1]] = category_count.get(row[1], 0) + 1
    print(Fore.MAGENTA + f"Total Items: {total_items}, Total Quantity: {total_quantity}")
    print(Fore.MAGENTA + "Category Summary:")
    for category, count in category_count.items():
        print(Fore.MAGENTA + f"  {category}: {count}")

# Function to alert on low stock items
def low_stock_alert(threshold=10):
    worksheet = load_worksheet()
    # (Code for checking and displaying low stock items)
    print(Fore.RED + "Low Stock Items:")
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[4] < threshold:
            print(Fore.RED + f"  {row[0]} (Quantity: {row[4]})")

# Function to warn about items nearing expiration
def expiration_warning(days=7):
    worksheet = load_worksheet()
    # (Code for checking and displaying items nearing expiration)
    today = datetime.now()
    warning_date = today + timedelta(days=days)
    print(Fore.YELLOW + "Items Nearing Expiration Date:")
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if isinstance(row[3], datetime) and today <= row[3] <= warning_date:
            print(Fore.YELLOW + f"  {row[0]} (Expires on: {row[3].strftime('%Y-%m-%d')})")

# Function to search items by category
def search_by_category():
    worksheet = load_worksheet()
    # (Code for taking category input and displaying matching items)
    category = input(Fore.WHITE + "Enter category to search (or type 'cancel' to exit): ")
    if category.lower() == 'cancel':
        return

    found = False
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[1] == category:
            print(Fore.GREEN + f"Item: {row[0]}, Category: {row[1]}, Price: {row[2]}, Expiration Date: {row[3].strftime('%Y-%m-%d') if isinstance(row[3], datetime) else 'N/A'}, Quantity: {row[4]}")
            found = True
    if not found:
        print(Fore.RED + f"No items found in category '{category}'.")

# Function to search for a specific item
def search_item():
    worksheet = load_worksheet()
    # (Code for taking item name input and displaying item details if found)
    name = input(Fore.WHITE + "Enter item name to search (or type 'cancel' to exit): ")
    if name.lower() == 'cancel':
        return

    try:
        found = False
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] == name:
                print(Fore.GREEN + f"Item: {row[0]}, Category: {row[1]}, Price: {row[2]}, Expiration Date: {row[3]}, Quantity: {row[4]}")
                found = True
                break
        if not found:
            print(Fore.RED + f"Item '{name}' not found.")
    except Exception as e:
        print(Fore.RED + f"An error occurred: {e}")

# Function for user authentication
def authenticate_user():
    valid_credentials = {
        'user1': 'password1',
        'user2': 'password2'
    }
    # Taking username and password input
    username = input(Fore.WHITE + "Enter username: ")
    password = input(Fore.WHITE + "Enter password: ")
    # (Input logic and authentication check)
    if username in valid_credentials and valid_credentials[username] == password:
        print(Fore.GREEN + "Authentication successful.")
        return True
    else:
        print(Fore.RED + "Invalid username or password.")
        return False
    
# Main function to run the inventory management system
def main():
    if not authenticate_user():
        print(Fore.RED + "Access denied. Exiting the program.")
        return

    print(Fore.CYAN + "\nWelcome to Supermarket Inventory Management System")

    while True:
        print(Fore.BLUE + "1. Display Inventory")
        print(Fore.BLUE + "2. Add Item")
        print(Fore.BLUE + "3. Update Item Details")
        print(Fore.BLUE + "4. Delete Item")
        print(Fore.BLUE + "5. Search for an Item")
        print(Fore.BLUE + "6. Search by Category")
        print(Fore.BLUE + "7. Inventory Summary")
        print(Fore.BLUE + "8. Low Stock Alert")
        print(Fore.BLUE + "9. Expiration Date Warning")
        print(Fore.BLUE + "10. Exit")
        choice = input(Fore.WHITE + "Enter your choice: ")

        if choice == '1':
            display_inventory()
        elif choice == '2':
            add_item()
        elif choice == '3':
            update_item()
        elif choice == '4':
            delete_item()
        elif choice == '5':
            search_item()
        elif choice == '6':
            search_by_category()
        elif choice == '7':
            inventory_summary()
        elif choice == '8':
            low_stock_alert()
        elif choice == '9':
            expiration_warning()
        elif choice == '10':
            print(Fore.RED + "Exiting the program. Goodbye!")
            break
        else:
            print(Fore.RED + "Invalid choice. Please try again.")

            # Display welcome message and menu options
        # (Code for displaying options and taking user choice)

        # Logic for executing the chosen option
        # (Code for the main loop handling user choices)    

# Entry point for the script
if __name__ == "__main__":
    main()
